#!/usr/bin/env python3
"""사업자등록번호 → bizno.net 법인정보 + 휴폐업 조회
Usage: python3 bizno_scrape.py [config.json]
"""
import json, os, re, sys, time
import openpyxl, requests
from bs4 import BeautifulSoup

def load_config(path="config.json"):
    with open(path) as f: return json.load(f)

HEADERS = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36'}

def clean_company_name(name):
    c = name.replace('（','(').replace('）',')')
    for pat in [r'^\s*\(주\)\s*', r'\s*\(주\)\s*$', r'^\s*주식회사\s*', r'\s*주식회사\s*$',
                r'^\s*\(유\)\s*', r'\s*\(유\)\s*$', r'^\s*유한회사\s*', r'\s*유한회사\s*$',
                r'^\s*유한책임회사\s*', r'\s*유한책임회사\s*$',
                r'^\s*사단법인\s*', r'^\s*재단법인\s*', r'^\s*법무법인\s*', r'^\s*합자회사\s*',
                r'\s*\([A-Za-z].*?\)\s*$', r'\(주\)']:
        c = re.sub(pat, '', c)
    return c.strip()

def format_pin(pin):
    pin = pin.replace('-', '')
    if len(pin) == 10:
        return f"{pin[:3]}-{pin[3:5]}-{pin[5:]}"
    return pin

def load_pins(cfg):
    wb = openpyxl.load_workbook(cfg['excel_path'], data_only=True)
    ws = wb[cfg['excel_sheet']]
    col = cfg.get('excel_pin_column', 6)
    seen, pins = set(), []
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=col).value
        if val:
            pin = str(val).strip().replace('-','')
            if pin and pin not in seen:
                seen.add(pin)
                pins.append(pin)
    return pins

def scrape_bizno(pin):
    query = pin.replace('-', '')
    resp = requests.get(f'https://bizno.net/?query={query}', headers=HEADERS, timeout=10)
    soup = BeautifulSoup(resp.text, 'html.parser')
    a = soup.select_one('a[href*="/article/"]')
    if not a:
        return {"pin": pin, "formatted_pin": format_pin(pin), "error": "no_result"}

    url = 'https://bizno.net' + a['href'] if a['href'].startswith('/') else a['href']
    title = a.get_text(strip=True)
    time.sleep(0.3)

    resp2 = requests.get(url, headers=HEADERS, timeout=10)
    text2 = resp2.text
    soup2 = BeautifulSoup(text2, 'html.parser')

    info = {"pin": pin, "formatted_pin": format_pin(pin)}
    for row in soup2.select('table tr'):
        th = row.select_one('th')
        td = row.select_one('td')
        if th and td:
            key = th.get_text(strip=True)
            val = td.get_text(strip=True)
            if key == '상호': info['company_name'] = val
            elif key == '대표자': info['representative'] = val
            elif key == '전화번호': info['phone'] = re.sub(r'\s*\(※.*?\)', '', val).strip()
            elif key == '사업자등록번호': info['biz_reg_display'] = val
            elif key == '법인등록번호': info['corp_reg_number'] = val
            elif key == '업태': info['business_type'] = val
            elif key == '업종': info['business_category'] = val
            elif key == '주소': info['address'] = val

    if 'company_name' not in info and title:
        info['company_name'] = title

    # 휴폐업 상태 확인
    match = re.search(r'사업자\s*현재\s*상태[^:]*:\s*(계속사업자|폐업자|휴업자)', text2)
    info['biz_status'] = match.group(1) if match else '확인불가'

    return info

def main():
    cfg_path = sys.argv[1] if len(sys.argv) > 1 else "config.json"
    cfg = load_config(cfg_path)

    pins = load_pins(cfg)
    print(f"대상 사업자등록번호: {len(pins)}건")

    cache_path = cfg.get('bizno_cache', './data/bizno_cache.json')
    output_path = cfg.get('bizno_results', './data/bizno_results.json')
    companies_path = cfg.get('companies_list', './data/iros_companies.json')

    os.makedirs(os.path.dirname(cache_path) or '.', exist_ok=True)

    cache = {}
    if os.path.exists(cache_path):
        with open(cache_path) as f: cache = json.load(f)
        print(f"캐시: {len(cache)}건")

    results, new_count, fail_count = [], 0, 0

    for i, pin in enumerate(pins):
        if pin in cache and 'error' not in cache[pin]:
            results.append(cache[pin])
            continue
        try:
            info = scrape_bizno(pin)
            cache[pin] = info
            results.append(info)
            name = info.get('company_name', '?')
            status = info.get('biz_status', '')
            print(f"  [{i+1}/{len(pins)}] {pin} -> {name} | {status}")
            new_count += 1
            time.sleep(1.5)
        except Exception as e:
            cache[pin] = {"pin": pin, "error": str(e)}
            results.append(cache[pin])
            fail_count += 1
            print(f"  [{i+1}/{len(pins)}] X {pin} -> {e}")
            time.sleep(2)

        if (new_count + fail_count) % 20 == 0 and (new_count + fail_count) > 0:
            with open(cache_path, 'w') as f: json.dump(cache, f, ensure_ascii=False, indent=2)

    with open(cache_path, 'w') as f: json.dump(cache, f, ensure_ascii=False, indent=2)
    with open(output_path, 'w') as f: json.dump(results, f, ensure_ascii=False, indent=2)

    # IROS용 정제 회사명 리스트
    companies = []
    for r in results:
        name = r.get('company_name', '')
        if name and r.get('biz_status') != '폐업자':
            companies.append(clean_company_name(name))
    with open(companies_path, 'w') as f: json.dump(companies, f, ensure_ascii=False, indent=2)

    print(f"\n총: {len(pins)}건, 신규: {new_count}, 실패: {fail_count}")
    print(f"정상 법인: {len(companies)}건 -> {companies_path}")

if __name__ == '__main__':
    main()
