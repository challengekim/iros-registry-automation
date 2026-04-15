#!/usr/bin/env python3
"""CDD 엑셀 생성
bizno 결과 + User ID 매핑 + 다운로드 상태를 종합하여 CDD 엑셀을 생성합니다.
Usage: python3 cdd_generate.py [config.json]
"""
import json, os, re, sys
from difflib import SequenceMatcher
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def load_config(path="config.json"):
    with open(path) as f:
        return json.load(f)


def clean_for_match(name):
    """매칭용 이름 정규화"""
    if not name:
        return ""
    s = re.sub(r'[^가-힣a-zA-Z0-9]', '', name)
    s = re.sub(r'주식회사|유한회사|유한책임회사|사단법인|재단법인|법무법인|합자회사', '', s)
    return s.strip()


def fuzzy_match_file(company_name, pdf_files):
    """회사명과 가장 유사한 PDF 파일 찾기"""
    if not company_name or not pdf_files:
        return None

    n_clean = clean_for_match(company_name)
    best_score, best_file = 0.0, None

    for pdf in pdf_files:
        stem = os.path.splitext(os.path.basename(pdf))[0]
        f_clean = clean_for_match(stem)

        # 완전 일치
        if n_clean == f_clean:
            return pdf

        # 포함 관계
        if n_clean and f_clean:
            if n_clean in f_clean or f_clean in n_clean:
                score = len(min(n_clean, f_clean, key=len)) / max(len(max(n_clean, f_clean, key=len)), 1)
                if score > best_score:
                    best_score, best_file = score, pdf

        # 유사도
        score = SequenceMatcher(None, n_clean, f_clean).ratio()
        if score > best_score:
            best_score, best_file = score, pdf

    return best_file if best_score > 0.5 else None


def load_user_ids(cfg):
    """MA 엑셀에서 User ID 매핑 로드
    - 사업자등록번호 탭: B=userID, C=pin
    - 추가요청 탭: J=userID, K=pin
    """
    excel_path = cfg.get('excel_path', '')
    mapping = {}  # pin(숫자만) -> userID

    if not excel_path or not os.path.exists(excel_path):
        return mapping

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)

        # 사업자등록번호 탭
        if '사업자등록번호' in wb.sheetnames:
            ws = wb['사업자등록번호']
            for row in range(2, ws.max_row + 1):
                uid = ws.cell(row=row, column=2).value
                pin = ws.cell(row=row, column=3).value
                if uid and pin:
                    pin_clean = re.sub(r'\D', '', str(pin))
                    if pin_clean:
                        mapping[pin_clean] = str(uid).strip()

        # 추가요청 탭
        sheet_name = cfg.get('excel_sheet', 'Sheet1')
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            uid_col = cfg.get('excel_userid_column', 10)
            pin_col = cfg.get('excel_pin_column', 6)
            for row in range(2, ws.max_row + 1):
                uid = ws.cell(row=row, column=uid_col).value
                pin = ws.cell(row=row, column=pin_col).value
                if uid and pin:
                    pin_clean = re.sub(r'\D', '', str(pin))
                    if pin_clean:
                        mapping[pin_clean] = str(uid).strip()

    except Exception as e:
        print(f"  [경고] User ID 로드 실패: {e}")

    return mapping


def get_pdf_files(save_dir):
    """저장 폴더에서 PDF 파일 목록"""
    save_dir = os.path.expanduser(save_dir)
    if not os.path.exists(save_dir):
        return []
    return [
        os.path.join(save_dir, f)
        for f in os.listdir(save_dir)
        if f.lower().endswith('.pdf')
    ]


def make_header_style():
    """헤더 셀 스타일"""
    font = Font(bold=True, color="FFFFFF", size=10)
    fill = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="FFFFFF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return font, fill, alignment, border


def make_data_style(row_idx, status):
    """데이터 셀 스타일 (상태별 배경색)"""
    alignment = Alignment(vertical="center", wrap_text=True)
    if status == "완료":
        fill_color = "E8F5E9" if row_idx % 2 == 0 else "F1F8E9"
    elif status == "실패":
        fill_color = "FFEBEE" if row_idx % 2 == 0 else "FFF3E0"
    else:
        fill_color = "F5F5F5" if row_idx % 2 == 0 else "FAFAFA"
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    return fill, alignment


def main():
    cfg_path = sys.argv[1] if len(sys.argv) > 1 else "config.json"
    cfg = load_config(cfg_path)

    bizno_path = cfg.get('bizno_results', './data/bizno_results.json')
    download_log_path = cfg.get('download_log', './logs/download_log.json')
    save_dir = cfg.get('save_dir', '~/Downloads/등기부등본')
    output_path = cfg.get('cdd_output', './output/CDD_고객정보.xlsx')

    # 데이터 로드
    print("데이터 로드 중...")
    with open(bizno_path) as f:
        bizno_results = json.load(f)

    download_log = {"completed": [], "failed": [], "skipped": []}
    if os.path.exists(download_log_path):
        with open(download_log_path) as f:
            download_log = json.load(f)

    pdf_files = get_pdf_files(save_dir)
    print(f"  PDF 파일: {len(pdf_files)}건")

    user_ids = load_user_ids(cfg)
    print(f"  User ID 매핑: {len(user_ids)}건")

    # 다운로드 완료된 상호 집합
    completed_sanghos = set(
        c.get('matched', c.get('sangho', ''))
        for c in download_log.get('completed', [])
    )

    # 각 bizno 항목에 대해 PDF 매칭 + User ID 조회
    rows_completed = []
    rows_failed = []
    rows_rest = []

    for r in bizno_results:
        if 'error' in r:
            continue

        company_name = r.get('company_name', '')
        pin_raw = re.sub(r'\D', '', r.get('pin', ''))
        corp_reg = r.get('corp_reg_number', '')
        phone = r.get('phone', '')
        address = r.get('address', '')
        representative = r.get('representative', '')
        business_type = r.get('business_type', '')
        biz_status = r.get('biz_status', '확인불가')
        formatted_pin = r.get('formatted_pin', r.get('pin', ''))

        # User ID 조회
        user_id = user_ids.get(pin_raw, '')

        # PDF 파일 매칭
        matched_file = fuzzy_match_file(company_name, pdf_files)

        if matched_file:
            dl_status = "완료"
            bigo = ""
        elif biz_status == "폐업자":
            dl_status = "실패"
            bigo = "폐업 법인"
        elif biz_status == "휴업자":
            dl_status = "실패"
            bigo = "휴업 법인"
        else:
            dl_status = "미완료"
            bigo = ""

        row = {
            "상호_한글": company_name,
            "상호_영문": "",
            "등록번호": corp_reg,
            "사업자등록번호": formatted_pin,
            "법인구분": "법인",
            "주소": address,
            "연락처": phone,
            "대표자": representative,
            "등기부_주소": "",
            "등기부_목적": "",
            "등기부_자본금": "",
            "등기부_발행주식": "",
            "등기부_주주": "",
            "등기부_임원": "",
            "등기부_비고": "",
            "user_id": user_id,
            "상태": dl_status,
            "비고": bigo,
            "pdf_file": matched_file or "",
            "biz_status": biz_status,
        }

        if dl_status == "완료":
            rows_completed.append(row)
        elif dl_status == "실패":
            rows_failed.append(row)
        else:
            rows_rest.append(row)

    # 정렬: 완료 -> 실패 -> 미완료
    all_rows = rows_completed + rows_failed + rows_rest
    print(f"  완료: {len(rows_completed)}, 실패: {len(rows_failed)}, 미완료: {len(rows_rest)}")

    # 엑셀 생성
    os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CDD 고객정보"

    # 헤더
    headers = [
        "A. 상호(한글)", "B. 상호(영문)", "C. 법인등록번호", "D. 사업자등록번호",
        "E. 법인구분", "F. 주소", "G. 연락처", "H. 대표자성명",
        "I. 등기_주소", "J. 등기_목적", "K. 등기_자본금",
        "L. 등기_발행주식", "M. 등기_주주", "N. 등기_임원", "O. 등기_비고",
        "P. User ID", "Q. 등기부등본 상태", "R. 비고",
    ]

    h_font, h_fill, h_align, h_border = make_header_style()
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = h_font
        cell.fill = h_fill
        cell.alignment = h_align
        cell.border = h_border

    ws.row_dimensions[1].height = 30

    # 열 너비 설정
    col_widths = [25, 20, 18, 16, 10, 40, 15, 15, 40, 30, 12, 12, 20, 20, 15, 15, 12, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # 데이터 행
    for row_idx, row in enumerate(all_rows, 2):
        status = row["상태"]
        fill, alignment = make_data_style(row_idx, status)

        values = [
            row["상호_한글"], row["상호_영문"], row["등록번호"], row["사업자등록번호"],
            row["법인구분"], row["주소"], row["연락처"], row["대표자"],
            row["등기부_주소"], row["등기부_목적"], row["등기부_자본금"],
            row["등기부_발행주식"], row["등기부_주주"], row["등기부_임원"], row["등기부_비고"],
            row["user_id"], row["상태"], row["비고"],
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            cell.alignment = alignment
            ws.row_dimensions[row_idx].height = 18

    # 틀 고정 (헤더 행)
    ws.freeze_panes = "A2"

    # 자동 필터
    ws.auto_filter.ref = f"A1:R{len(all_rows)+1}"

    wb.save(output_path)
    print(f"\nCDD 엑셀 저장 완료: {output_path}")
    print(f"총 {len(all_rows)}건 (완료:{len(rows_completed)} / 실패:{len(rows_failed)} / 미완료:{len(rows_rest)})")


if __name__ == "__main__":
    main()
